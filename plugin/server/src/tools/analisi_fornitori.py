"""MCP tools for supplier-ledger privacy screening (analisi mastrino fornitori).

TRIGGER: usare quando l'utente deve analizzare il mastrino fornitori di un cliente
ai fini GDPR (individuare i responsabili ex art. 28 da nominare), verificare una
P.IVA sul VIES, o produrre il report Excel standard dell'analisi fornitori.
"""

import os
import re
import tempfile
import uuid
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.lib.vies import check_vat, checksum_partita_iva
from src.server import mcp


QUALIFICAZIONI = {"responsabile", "titolare_autonomo", "fuori_perimetro"}
CONFIDENZE = {"alto", "medio", "basso"}
PROBABILITA = {"alta", "media", "bassa"}
DPA_VALORI = {"si", "no", "da_verificare"}

_CAMPI_OBBLIGATORI = ("denominazione_mastrino", "qualificazione", "motivazione", "confidenza")


def _valida_fornitori(fornitori) -> list[str]:
    """Collect-all validation of canonical supplier records (1-based row indexes)."""
    if not isinstance(fornitori, list):
        return ["'fornitori' deve essere una lista di oggetti"]
    if not fornitori:
        return ["'fornitori' è una lista vuota: nessun fornitore da riportare"]

    errori: list[str] = []
    for i, riga in enumerate(fornitori, start=1):
        if not isinstance(riga, dict):
            errori.append(f"riga {i}: non è un oggetto")
            continue
        for campo in _CAMPI_OBBLIGATORI:
            valore = riga.get(campo)
            if not isinstance(valore, str) or not valore.strip():
                errori.append(f"riga {i}: campo obbligatorio '{campo}' mancante o vuoto")
        qualificazione = riga.get("qualificazione")
        if isinstance(qualificazione, str) and qualificazione and qualificazione not in QUALIFICAZIONI:
            errori.append(
                f"riga {i}: 'qualificazione' non valida ({qualificazione!r}); ammessi: {sorted(QUALIFICAZIONI)}"
            )
        confidenza = riga.get("confidenza")
        if isinstance(confidenza, str) and confidenza and confidenza not in CONFIDENZE:
            errori.append(f"riga {i}: 'confidenza' non valida ({confidenza!r}); ammessi: {sorted(CONFIDENZE)}")

        probabilita = riga.get("probabilita_responsabile")
        dpa = riga.get("dpa_proprio")
        if qualificazione == "responsabile":
            if not isinstance(probabilita, str) or probabilita not in PROBABILITA:
                errori.append(
                    f"riga {i}: 'probabilita_responsabile' obbligatoria per i responsabili; ammessi: {sorted(PROBABILITA)}"
                )
            if not isinstance(dpa, str) or dpa not in DPA_VALORI:
                errori.append(
                    f"riga {i}: 'dpa_proprio' obbligatorio per i responsabili; ammessi: {sorted(DPA_VALORI)}"
                )
        elif isinstance(qualificazione, str) and qualificazione in QUALIFICAZIONI:
            if probabilita is not None:
                errori.append(
                    f"riga {i}: 'probabilita_responsabile' presente ma la qualificazione non è 'responsabile'"
                )
            if dpa is not None:
                errori.append(f"riga {i}: 'dpa_proprio' presente ma la qualificazione non è 'responsabile'")

        fonti = riga.get("fonti", [])
        if fonti is not None and not isinstance(fonti, list):
            errori.append(f"riga {i}: 'fonti' deve essere una lista di URL")
        elif isinstance(fonti, list) and not all(isinstance(elem, str) for elem in fonti):
            errori.append(f"riga {i}: 'fonti' deve contenere solo stringhe (URL)")

        for campo in ("piva_cf", "fonte_piva", "attivita", "categorie_dati", "note"):
            valore_campo = riga.get(campo)
            if valore_campo is not None and not isinstance(valore_campo, str):
                errori.append(f"riga {i}: campo '{campo}' deve essere una stringa")
    return errori


_OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "mcp-legal-it")

_HEADERS = (
    "Denominazione (da mastrino)",
    "P.IVA / CF",
    "Attività / servizi",
    "Categorie di dati presumibilmente trattate",
    "Qualificazione ipotizzata",
    "Motivazione sintetica",
    "Probabilità che tratti dati come responsabile",
    "DPA proprio del fornitore disponibile?",
    "Confidenza dell'identificazione",
    "Fonte (URL)",
    "Note / flag",
)
_COL_WIDTHS = (30, 16, 34, 34, 20, 40, 16, 16, 14, 32, 40)

_LABEL_QUALIFICAZIONE = {
    "responsabile": "Responsabile del trattamento",
    "titolare_autonomo": "Titolare autonomo",
    "fuori_perimetro": "Fuori perimetro privacy",
}
_LABEL_PROBABILITA = {"alta": "Alta", "media": "Media", "bassa": "Bassa"}
_LABEL_DPA = {"si": "Sì", "no": "No", "da_verificare": "Da verificare"}
_LABEL_CONFIDENZA = {"alto": "Alto", "medio": "Medio", "basso": "Basso"}

_DISCLAIMER = (
    "Analisi automatica di primo livello, da validare con il cliente e con i contratti. "
    "Ove manchi la P.IVA alcune identificazioni sono incerte: verificare manualmente le "
    "voci con Confidenza \"Basso\" e i flag \"controverso\" nelle Note."
)

_ORDINE_DPA = {"no": 0, "da_verificare": 1, "si": 2}


def _neutralizza_formula(cella) -> None:
    """openpyxl treats a string starting with '=' as a live formula. Ledger data
    (denominazione_mastrino, note, etc.) is third-party-controlled text, so force
    such cells back to a plain string type to prevent formula injection on open."""
    if isinstance(cella.value, str) and cella.value.startswith("="):
        cella.data_type = "s"


def _chiave_ordinamento(riga: dict) -> tuple:
    qualificazione = riga["qualificazione"]
    if qualificazione == "responsabile":
        gruppo = _ORDINE_DPA[riga["dpa_proprio"]]
    elif qualificazione == "titolare_autonomo":
        gruppo = 3
    else:
        gruppo = 4
    return (gruppo, riga["denominazione_mastrino"].upper())


def _ordina(fornitori: list[dict]) -> list[dict]:
    """No-DPA responsabili first (they need a nomina), then the rest; A-Z within groups."""
    return sorted(fornitori, key=_chiave_ordinamento)


def _sanitize_filename(name: str) -> str:
    name = name.lower().replace(" ", "_")
    name = re.sub(r"[^a-z0-9_\-]", "", name)
    return name[:50] or "cliente"


def _scrivi_avvertenze(ws, cliente: str, data_analisi: str, file_sorgente: str, fornitori: list[dict]) -> None:
    conte = {"responsabile": 0, "titolare_autonomo": 0, "fuori_perimetro": 0}
    nomine = 0
    for riga in fornitori:
        conte[riga["qualificazione"]] += 1
        if riga["qualificazione"] == "responsabile" and riga["dpa_proprio"] == "no":
            nomine += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    righe = [
        ("Analisi privacy fornitori", ""),
        ("Cliente (titolare)", cliente),
        ("Data analisi", data_analisi),
        ("File sorgente", file_sorgente or "—"),
        ("Totale fornitori analizzati", len(fornitori)),
        ("Responsabili del trattamento", conte["responsabile"]),
        ("— di cui senza DPA proprio (nomina da predisporre)", nomine),
        ("Titolari autonomi", conte["titolare_autonomo"]),
        ("Fuori perimetro privacy", conte["fuori_perimetro"]),
        ("", ""),
        ("AVVERTENZE", _DISCLAIMER),
    ]
    for r, (etichetta, valore) in enumerate(righe, start=1):
        ws.cell(row=r, column=1, value=etichetta).font = Font(bold=True)
        cella = ws.cell(row=r, column=2, value=valore)
        _neutralizza_formula(cella)
        cella.alignment = Alignment(wrap_text=True, vertical="top")


def _scrivi_analisi(ws, fornitori: list[dict]) -> None:
    intestazione_font = Font(bold=True, color="FFFFFF")
    intestazione_fill = PatternFill("solid", fgColor="4472C4")
    for col, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cella = ws.cell(row=1, column=col, value=header)
        cella.font = intestazione_font
        cella.fill = intestazione_fill
        cella.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    for r, riga in enumerate(fornitori, start=2):
        responsabile = riga["qualificazione"] == "responsabile"
        valori = (
            riga["denominazione_mastrino"],
            riga.get("piva_cf") or "",
            riga.get("attivita") or "",
            riga.get("categorie_dati") or "",
            _LABEL_QUALIFICAZIONE[riga["qualificazione"]],
            riga["motivazione"],
            _LABEL_PROBABILITA[riga["probabilita_responsabile"]] if responsabile else "—",
            _LABEL_DPA[riga["dpa_proprio"]] if responsabile else "—",
            _LABEL_CONFIDENZA[riga["confidenza"]],
            "\n".join(riga.get("fonti") or []),
            riga.get("note") or "",
        )
        for col, valore in enumerate(valori, start=1):
            cella = ws.cell(row=r, column=col, value=valore)
            _neutralizza_formula(cella)
            cella.alignment = Alignment(wrap_text=True, vertical="top")


@mcp.tool(tags={"privacy", "utility"})
def genera_report_fornitori(
    fornitori: list[dict],
    cliente: str,
    data_analisi: str = "",
    file_sorgente: str = "",
    nome_file: str = "",
) -> str:
    """Genera l'Excel standard dell'analisi privacy del mastrino fornitori (foglio Avvertenze + 11 colonne).

    Riceve le righe già classificate nel formato canonico dell'analisi fornitori
    (vedi skill analisi-fornitori) e produce SEMPRE lo stesso layout: responsabili
    senza DPA proprio in cima (sono le nomine da predisporre), poi gli altri
    responsabili, i titolari autonomi e i fuori perimetro. Valida ogni riga e in
    caso di errori li restituisce tutti insieme senza scrivere il file.

    Vigenza: art. 28 GDPR (nomina responsabile); art. 4 GDPR (definizioni).
    Precisione: ESATTO per il layout; il contenuto riflette l'analisi ricevuta.

    Args:
        fornitori: Lista di record canonici (denominazione_mastrino, qualificazione,
            motivazione, confidenza obbligatori; probabilita_responsabile e
            dpa_proprio solo per i responsabili; piva_cf, attivita, categorie_dati,
            fonti, note opzionali)
        cliente: Denominazione del titolare (il cliente dello studio)
        data_analisi: Data dell'analisi in formato gg/mm/aaaa (default: oggi)
        file_sorgente: Nome del file mastrino analizzato (mostrato in Avvertenze)
        nome_file: Nome file di output personalizzato (default generato dal cliente)
    """
    errori = _valida_fornitori(fornitori)
    if errori:
        return "Errore di validazione: " + "; ".join(errori)
    if not cliente or not cliente.strip():
        return "Errore di validazione: 'cliente' è obbligatorio"

    data_analisi = data_analisi.strip() or date.today().strftime("%d/%m/%Y")
    ordinate = _ordina(fornitori)

    wb = Workbook()
    ws_avvertenze = wb.active
    ws_avvertenze.title = "Avvertenze"
    _scrivi_avvertenze(ws_avvertenze, cliente.strip(), data_analisi, file_sorgente.strip(), ordinate)
    _scrivi_analisi(wb.create_sheet("Analisi fornitori"), ordinate)

    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    nome_file = nome_file.strip()
    if not nome_file:
        nome_file = f"analisi_fornitori_{_sanitize_filename(cliente)}_{uuid.uuid4().hex[:8]}.xlsx"
    elif not nome_file.lower().endswith(".xlsx"):
        nome_file = nome_file + ".xlsx"
    filepath = os.path.join(_OUTPUT_DIR, os.path.basename(nome_file))
    wb.save(filepath)
    size_kb = round(os.path.getsize(filepath) / 1024, 1)
    return f"File salvato: {filepath} ({size_kb} KB)"


@mcp.tool(tags={"privacy", "utility"})
async def verifica_partita_iva_vies(partita_iva: str, codice_paese: str = "IT") -> dict:
    """Verifica una partita IVA sul VIES (servizio UE gratuito): validità e, se disponibili, denominazione e indirizzo registrati.

    Per le P.IVA italiane esegue prima il checksum locale: se fallisce, il VIES non
    viene interrogato. Usare per agganciare con certezza l'identità di un fornitore
    (es. durante l'analisi del mastrino fornitori). `disponibile: false` significa
    VIES/stato membro momentaneamente non raggiungibile: procedere con la sola
    ricerca web e annotarlo.

    Vigenza: Regolamento (UE) 904/2010 (cooperazione amministrativa IVA); DPR 633/1972 art. 35 per il checksum.
    Precisione: ESATTO per validità; denominazione/indirizzo dipendono dai dati forniti dallo stato membro.

    Args:
        partita_iva: Numero IVA senza prefisso paese (per l'Italia: 11 cifre)
        codice_paese: Codice ISO dello stato membro (default "IT")
    """
    piva = partita_iva.strip().replace(" ", "")
    paese = codice_paese.strip().upper() or "IT"
    checksum: bool | None = checksum_partita_iva(piva) if paese == "IT" else None

    if checksum is False:
        return {
            "partita_iva": piva,
            "codice_paese": paese,
            "checksum_valido": False,
            "disponibile": None,
            "valido": False,
            "denominazione": None,
            "indirizzo": None,
            "errore": "checksum non valido — VIES non interrogato",
        }

    esito = await check_vat(piva, paese)
    return {
        "partita_iva": piva,
        "codice_paese": paese,
        "checksum_valido": checksum,
        **esito,
    }
