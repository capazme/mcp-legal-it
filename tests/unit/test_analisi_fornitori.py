"""Unit tests for genera_report_fornitori: validation and xlsx rendering."""

import importlib

from src.tools.analisi_fornitori import _valida_fornitori


def _tool(fn_name: str):
    mod = importlib.import_module("src.tools.analisi_fornitori")
    fn = getattr(mod, fn_name)
    return fn.fn if hasattr(fn, "fn") else fn


def _riga_ok(**overrides) -> dict:
    """A fully valid 'responsabile' canonical record; override per-test."""
    base = {
        "denominazione_mastrino": "ACME CLOUD SRL",
        "piva_cf": "01234567890",
        "fonte_piva": "mastrino",
        "attivita": "Hosting e SaaS gestionale",
        "categorie_dati": "Dati di clienti/utenti del titolare",
        "qualificazione": "responsabile",
        "motivazione": "SaaS che tratta dati per conto del titolare",
        "probabilita_responsabile": "alta",
        "dpa_proprio": "no",
        "confidenza": "medio",
        "fonti": ["https://esempio.it"],
        "note": "",
    }
    base.update(overrides)
    return {k: v for k, v in base.items() if v is not None or k in overrides}


class TestValidazione:
    def test_lista_valida(self):
        assert _valida_fornitori([_riga_ok()]) == []

    def test_lista_vuota(self):
        errs = _valida_fornitori([])
        assert len(errs) == 1 and "vuota" in errs[0]

    def test_non_lista(self):
        errs = _valida_fornitori("non una lista")
        assert len(errs) == 1

    def test_riga_non_dict(self):
        errs = _valida_fornitori(["stringa"])
        assert errs and errs[0].startswith("riga 1:")

    def test_campi_obbligatori_mancanti(self):
        errs = _valida_fornitori([_riga_ok(denominazione_mastrino="", motivazione=None)])
        joined = " | ".join(errs)
        assert "denominazione_mastrino" in joined and "motivazione" in joined

    def test_enum_qualificazione(self):
        errs = _valida_fornitori([_riga_ok(qualificazione="RESPONSABILE")])
        assert errs and "qualificazione" in errs[0]

    def test_enum_confidenza(self):
        errs = _valida_fornitori([_riga_ok(confidenza="altissimo")])
        assert errs and "confidenza" in errs[0]

    def test_responsabile_richiede_probabilita_e_dpa(self):
        errs = _valida_fornitori([_riga_ok(probabilita_responsabile=None, dpa_proprio=None)])
        joined = " | ".join(errs)
        assert "probabilita_responsabile" in joined and "dpa_proprio" in joined

    def test_non_responsabile_vieta_campi_responsabile(self):
        riga = _riga_ok(qualificazione="titolare_autonomo")
        errs = _valida_fornitori([riga])
        joined = " | ".join(errs)
        assert "probabilita_responsabile" in joined and "dpa_proprio" in joined

    def test_titolare_autonomo_valido(self):
        riga = _riga_ok(
            qualificazione="titolare_autonomo",
            probabilita_responsabile=None,
            dpa_proprio=None,
        )
        riga.pop("probabilita_responsabile")
        riga.pop("dpa_proprio")
        assert _valida_fornitori([riga]) == []

    def test_fonti_non_lista(self):
        errs = _valida_fornitori([_riga_ok(fonti="https://esempio.it")])
        assert errs and "fonti" in errs[0]

    def test_indici_multipli(self):
        errs = _valida_fornitori([_riga_ok(), _riga_ok(confidenza="x"), _riga_ok(qualificazione="y")])
        assert any(e.startswith("riga 2:") for e in errs)
        assert any(e.startswith("riga 3:") for e in errs)

    def test_qualificazione_unhashable(self):
        """qualificazione as list → 'campo obbligatorio' error, no TypeError."""
        errs = _valida_fornitori([_riga_ok(qualificazione=["responsabile"])])
        assert errs
        assert any("qualificazione" in e and "mancante o vuoto" in e for e in errs)

    def test_probabilita_dpa_unhashable(self):
        """probabilita and dpa as unhashable types → errors, no TypeError."""
        errs = _valida_fornitori([_riga_ok(probabilita_responsabile=["alta"], dpa_proprio={"si"})])
        assert errs
        assert any("probabilita_responsabile" in e for e in errs)
        assert any("dpa_proprio" in e for e in errs)

    def test_fonti_con_elemento_non_stringa(self):
        """A non-str item in 'fonti' must be a collected error, not a raw TypeError at render time."""
        errs = _valida_fornitori([_riga_ok(fonti=["https://esempio.it", 42])])
        assert errs
        assert any("riga 1" in e and "fonti" in e and "stringhe" in e for e in errs)

    def test_note_dict_e_campo_non_stringa(self):
        """A dict in 'note' (or any optional scalar field) must be a collected error."""
        errs = _valida_fornitori([_riga_ok(note={"flag": "controverso"})])
        assert errs
        assert any("riga 1" in e and "note" in e for e in errs)


# ---------------------------------------------------------------------------
# Rendering xlsx
# ---------------------------------------------------------------------------

from openpyxl import load_workbook

from src.tools.analisi_fornitori import _ordina

_HEADER_ATTESO = [
    "Denominazione (da mastrino)",
    "P.IVA / CF",
    "Attività / servizi",
    "Categorie di dati presumibilmente trattate",
    "Qualificazione ipotizzata",
    "Motivazione sintetica",
    "Probabilità che tratti dati come responsabile",
    "DPA proprio del fornitore disponibile?",
    "Confidenza dell'identificazione",
    "Fonte (URL)",
    "Note / flag",
]


def _fuori(nome="CARTOLERIA ROSSI"):
    riga = _riga_ok(
        denominazione_mastrino=nome,
        qualificazione="fuori_perimetro",
        motivazione="Fornitore di soli beni",
    )
    riga.pop("probabilita_responsabile")
    riga.pop("dpa_proprio")
    return riga


def _titolare(nome="STUDIO BIANCHI COMMERCIALISTI"):
    riga = _riga_ok(
        denominazione_mastrino=nome,
        qualificazione="titolare_autonomo",
        motivazione="Determina autonomamente finalità e mezzi",
    )
    riga.pop("probabilita_responsabile")
    riga.pop("dpa_proprio")
    return riga


class TestOrdina:
    def test_ordinamento_gruppi_e_alfabetico(self):
        righe = [
            _fuori("ZETA CANCELLERIA"),
            _titolare(),
            _riga_ok(denominazione_mastrino="B-CLOUD", dpa_proprio="si"),
            _riga_ok(denominazione_mastrino="A-CLOUD", dpa_proprio="no"),
            _riga_ok(denominazione_mastrino="C-CLOUD", dpa_proprio="da_verificare"),
            _riga_ok(denominazione_mastrino="AA-CLOUD", dpa_proprio="no"),
        ]
        ordinate = [r["denominazione_mastrino"] for r in _ordina(righe)]
        assert ordinate == [
            "A-CLOUD", "AA-CLOUD",              # responsabili dpa=no, alfabetico
            "C-CLOUD",                            # dpa=da_verificare
            "B-CLOUD",                            # dpa=si
            "STUDIO BIANCHI COMMERCIALISTI",      # titolari autonomi
            "ZETA CANCELLERIA",                   # fuori perimetro
        ]


class TestGeneraReport:
    def test_errore_validazione_non_scrive_file(self, tmp_path, monkeypatch):
        monkeypatch.setattr("src.tools.analisi_fornitori._OUTPUT_DIR", str(tmp_path))
        out = _tool("genera_report_fornitori")(
            fornitori=[_riga_ok(confidenza="x")], cliente="Cliente Srl"
        )
        assert out.startswith("Errore di validazione: riga 1:")
        assert list(tmp_path.iterdir()) == []

    def test_file_generato_struttura(self, tmp_path, monkeypatch):
        monkeypatch.setattr("src.tools.analisi_fornitori._OUTPUT_DIR", str(tmp_path))
        out = _tool("genera_report_fornitori")(
            fornitori=[_riga_ok(), _titolare(), _fuori()],
            cliente="Cliente Srl",
            data_analisi="30/07/2026",
            file_sorgente="mastrino.xlsx",
        )
        assert out.startswith("File salvato: ")
        files = list(tmp_path.glob("analisi_fornitori_cliente_srl_*.xlsx"))
        assert len(files) == 1

        wb = load_workbook(files[0])
        assert wb.sheetnames == ["Avvertenze", "Analisi fornitori"]

        ws = wb["Analisi fornitori"]
        header = [c.value for c in ws[1]]
        assert header == _HEADER_ATTESO
        assert ws.freeze_panes == "A2"

        prima_riga = [c.value for c in ws[2]]
        assert prima_riga[0] == "ACME CLOUD SRL"
        assert prima_riga[4] == "Responsabile del trattamento"
        assert prima_riga[6] == "Alta"
        assert prima_riga[7] == "No"
        assert prima_riga[8] == "Medio"
        assert prima_riga[9] == "https://esempio.it"

        riga_titolare = [c.value for c in ws[3]]
        assert riga_titolare[4] == "Titolare autonomo"
        assert riga_titolare[6] == "—"
        assert riga_titolare[7] == "—"

        riga_fuori = [c.value for c in ws[4]]
        assert riga_fuori[4] == "Fuori perimetro privacy"

    def test_avvertenze_contenuto(self, tmp_path, monkeypatch):
        monkeypatch.setattr("src.tools.analisi_fornitori._OUTPUT_DIR", str(tmp_path))
        _tool("genera_report_fornitori")(
            fornitori=[_riga_ok(), _titolare(), _fuori()],
            cliente="Cliente Srl",
            data_analisi="30/07/2026",
            file_sorgente="mastrino.xlsx",
        )
        wb = load_workbook(next(tmp_path.glob("*.xlsx")))
        ws = wb["Avvertenze"]
        testo = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "Cliente Srl" in testo
        assert "30/07/2026" in testo
        assert "mastrino.xlsx" in testo
        assert "validare" in testo               # disclaimer
        assert "Basso" in testo                  # review warning

        riga_totale = next(row for row in ws.iter_rows() if row[0].value == "Totale fornitori analizzati")
        assert riga_totale[1].value == 3

    def test_fonti_multiple_su_piu_righe(self, tmp_path, monkeypatch):
        monkeypatch.setattr("src.tools.analisi_fornitori._OUTPUT_DIR", str(tmp_path))
        _tool("genera_report_fornitori")(
            fornitori=[_riga_ok(fonti=["https://a.it", "https://b.it"])],
            cliente="X",
        )
        wb = load_workbook(next(tmp_path.glob("*.xlsx")))
        cella = wb["Analisi fornitori"].cell(row=2, column=10).value
        assert cella == "https://a.it\nhttps://b.it"

    def test_denominazione_formula_neutralizzata(self, tmp_path, monkeypatch):
        """A ledger-supplied denominazione starting with '=' must never become a live
        Excel formula — it must be stored as a literal string cell."""
        monkeypatch.setattr("src.tools.analisi_fornitori._OUTPUT_DIR", str(tmp_path))
        formula_payload = '=HYPERLINK("http://evil","x")'
        _tool("genera_report_fornitori")(
            fornitori=[_riga_ok(denominazione_mastrino=formula_payload)],
            cliente="X",
        )
        wb = load_workbook(next(tmp_path.glob("*.xlsx")))
        cella = wb["Analisi fornitori"].cell(row=2, column=1)
        assert cella.data_type != "f"
        assert cella.value == formula_payload
